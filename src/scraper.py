import requests
import feedparser
from bs4 import BeautifulSoup
import re
import openpyxl
from openpyxl import Workbook
import argparse
import logging # Added for logging

# Global logger instance (can be configured in main)
# However, best practice is to get logger in each module: logger = logging.getLogger(__name__)
# For a single file script, a global one configured in main is okay.
logger = logging.getLogger(__name__)

AWS_RSS_URL = "https://aws.amazon.com/about-aws/whats-new/recent/feed/"
AZURE_RSS_URL = "https://www.microsoft.com/releasecommunications/api/v2/azure/rss"

VIEW_TEXT_WEBSITE_AZURE_CONTENT = """\
<?xml version="1.0" encoding="utf-8"?><rss xmlns:a10="http://www.w3.org/2005/Atom" version="2.0"><channel><title>Azure service updates</title><link>https://www.microsoft.com/releasecommunications/api/v2/Azure/rss</link><description>Azure service updates</description><lastBuildDate>Fri, 30 May 2025 17:45:01 Z</lastBuildDate><ttl>1440</ttl><item><guid isPermaLink="false">495605</guid><link>https://azure.microsoft.com/en-us/updates/launched-generally-available-azure-quota-groups/</link><category>Launched</category><category>Compute</category><category>Virtual Machines</category><category>Features</category><title>[Launched] Generally Available: Azure Quota Groups</title><description>Azure Quota Groups is now generally available...</description><pubDate>Fri, 30 May 2025 17:45:01 Z</pubDate><a10:updated>2025-05-30T17:45:01Z</a10:updated></item><item><guid isPermaLink="false">492962</guid><link>https://azure.microsoft.com/en-us/updates/retirement-language-understanding-luis-retirement-extended-to-october-31-2025/</link><category>AI + machine learning</category><category>Mobile</category><category>Azure AI Services</category><category>Language Understanding (LUIS)</category><category>Retirements</category><title>Retirement: Language Understanding (LUIS) retirement extended to October 31, 2025</title><description>In September 2022, we announced that Language Understanding (LUIS) would be retired...</description><pubDate>Fri, 30 May 2025 11:15:51 Z</pubDate><a10:updated>2025-05-30T11:15:51Z</a10:updated></item><item><guid isPermaLink="false">490047</guid><link>https://azure.microsoft.com/en-us/updates/microsoft-fabric-runtime-1-2-retirement/</link><category>Analytics</category><title>Retirement: Microsoft Fabric Runtime 1.2</title><description>End of Support for Microsoft Fabric Runtime 1.2 has been announced.</description><pubDate>Fri, 30 May 2025 11:15:50 Z</pubDate><a10:updated>2025-05-30T11:15:50Z</a10:updated></item></channel></rss>
"""

KNOWN_AWS_PRODUCTS = {
    "EC2", "S3", "RDS", "VPC", "Lambda", "DynamoDB", "CloudFormation", "CloudWatch",
    "SNS", "SQS", "EKS", "ECS", "IAM", "Route 53", "Glacier", "Kinesis",
    "Redshift", "SageMaker", "Aurora", "AppSync", "Amplify", "Cognito", "Fargate"
}

def extract_product_names(title):
    if not title: return []
    products = set()
    words = re.findall(r'\b\w+\b', title)
    title_lower = title.lower()
    for product in KNOWN_AWS_PRODUCTS:
        if product.lower() in title_lower: products.add(product)
    for i in range(len(words) - 1):
        if words[i] in ("Amazon", "AWS") and words[i+1].istitle(): products.add(f"{words[i]} {words[i+1]}")
        if words[i].isupper() and 2 <= len(words[i]) <= 6 and words[i] not in ["NEW", "NOW", "AND", "FOR", "THE", "WITH"]:
            is_known_part = any(words[i] in known.split() for known in KNOWN_AWS_PRODUCTS)
            if not is_known_part: products.add(words[i])
    final_products = set(products)
    for p1 in products:
        for p2 in products:
            if p1 != p2 and p1 in p2 and p1 in final_products: final_products.remove(p1); break
    return list(final_products) if final_products else ["N/A"]

def scrape_aws_updates(limit=None):
    logger.info("Starting AWS updates scraping process...")
    aws_data_list = []
    try:
        response = requests.get(AWS_RSS_URL, timeout=10)
        response.raise_for_status()
        logger.info(f"Successfully fetched AWS RSS feed from {AWS_RSS_URL}")
    except requests.exceptions.RequestException as e:
        logger.error(f"Error fetching AWS RSS feed from {AWS_RSS_URL}: {e}")
        return aws_data_list

    feed = feedparser.parse(response.content)
    if feed.bozo:
        logger.warning(f"AWS Feed may be malformed. Bozo reason: {feed.bozo_exception}")
    if not feed.entries:
        logger.warning("No entries found in AWS RSS feed.")
        return aws_data_list

    entries_to_process = feed.entries
    if limit is not None and limit > 0:
        logger.info(f"Limiting AWS entries to first {limit}.")
        entries_to_process = feed.entries[:limit]

    for i, entry in enumerate(entries_to_process):
        try:
            # logger.debug(f"Processing AWS entry {i+1}/{len(entries_to_process)}: {entry.get('title', 'N/A')}")
            item_data = {"Title": entry.get("title", "N/A"), "URL": entry.get("link", "N/A"), "Date Posted": "N/A", "Description": "N/A", "Links": [], "Product Name": "N/A"}
            published_parsed = entry.get("published_parsed")
            if published_parsed: item_data["Date Posted"] = f"{published_parsed.tm_year}-{published_parsed.tm_mon:02d}-{published_parsed.tm_mday:02d}"
            item_data["Product Name"] = ", ".join(extract_product_names(item_data["Title"]))

            if item_data["URL"] == "N/A" or not item_data["URL"].startswith("http"):
                logger.warning(f"Skipping content scraping for AWS item '{item_data['Title']}' due to invalid URL: {item_data['URL']}")
                aws_data_list.append(item_data); continue

            # logger.debug(f"Fetching page content for AWS item: {item_data['URL']}")
            page_response = requests.get(item_data["URL"], timeout=15)
            page_response.raise_for_status()
            soup = BeautifulSoup(page_response.content, 'html.parser')
            main_content = soup.find('div', class_='aws-text-box') or soup.find('section', class_='aws-text-box') or soup.find('main', attrs={'role': 'main'}) or soup.find('article') or soup.body

            if main_content:
                paragraphs = main_content.find_all('p')
                desc_text = "\n".join(p.get_text(separator=' ', strip=True) for p in paragraphs) if paragraphs else main_content.get_text(separator=' ', strip=True)
                item_data["Description"] = (desc_text[:1020] + '...') if len(desc_text) > 1023 else desc_text
                links_on_page = [a['href'] for a in main_content.find_all('a', href=True)]
                item_data["Links"] = ", ".join(links_on_page[:10])
            else:
                item_data["Description"] = "Main description area not found."
                logger.warning(f"Could not find main description area for AWS URL: {item_data['URL']}")
        except requests.exceptions.RequestException as e_req:
            logger.error(f"Request error while processing AWS item '{entry.get('title', 'N/A')}' URL {item_data.get('URL', 'N/A')}: {e_req}")
            item_data["Description"] = f"Error fetching page: {e_req}" # Store error in description
        except Exception as e_gen:
            logger.error(f"Unexpected error while processing AWS item '{entry.get('title', 'N/A')}': {e_gen}", exc_info=True)
            item_data["Description"] = f"Unexpected error: {e_gen}" # Store error

        aws_data_list.append(item_data)
    logger.info(f"Finished processing {len(aws_data_list)} AWS entries.")
    return aws_data_list

def scrape_azure_updates(limit=None):
    logger.info("Starting Azure updates scraping process...")
    azure_data_list = []
    response = None
    try:
        response = requests.get(AZURE_RSS_URL, timeout=10)
        response.raise_for_status()
        logger.info(f"Successfully fetched Azure RSS feed from {AZURE_RSS_URL}")
    except requests.exceptions.RequestException as e:
        logger.error(f"Error fetching Azure RSS feed via requests from {AZURE_RSS_URL}: {e}")
        pass # Will try fallback to VIEW_TEXT_WEBSITE_AZURE_CONTENT

    decoded_content_requests = ""
    if response:
        try: decoded_content_requests = response.content.decode('utf-8-sig').strip()
        except UnicodeDecodeError as ude:
            logger.warning(f"UnicodeDecodeError for Azure feed from requests: {ude}. Trying response.text.")
            try: decoded_content_requests = response.text.strip()
            except Exception as e_text: logger.error(f"Could not get text content from Azure response: {e_text}"); decoded_content_requests = ""

    feed_parsed_successfully_fp = False
    feed_fp = None
    if decoded_content_requests:
        try:
            # logger.debug("Attempting to parse Azure feed with feedparser from requests content...")
            feed_fp = feedparser.parse(decoded_content_requests)
            if feed_fp.bozo:
                logger.warning(f"Feedparser (on Azure requests content): Feed may be malformed. Bozo reason: {feed_fp.bozo_exception}")
            if feed_fp.entries:
                # logger.info(f"Feedparser (on Azure requests content) found {len(feed_fp.entries)} entries.")
                feed_parsed_successfully_fp = True
            elif not feed_fp.bozo: logger.info("Feedparser (on Azure requests content) found no entries, but feed appears well-formed.")
            else: logger.warning("Feedparser (on Azure requests content) found no entries and feed is marked as not well-formed (bozo).")
        except Exception as e_fp: logger.error(f"Critical error during feedparser.parse() (Azure requests content): {e_fp}", exc_info=True)

    if not (feed_parsed_successfully_fp and feed_fp and feed_fp.entries):
        logger.info("Feedparser failed to get Azure entries or feed was empty/bozo. Attempting fallback with BeautifulSoup on VIEW_TEXT_WEBSITE_AZURE_CONTENT.")
        current_content_source = VIEW_TEXT_WEBSITE_AZURE_CONTENT
        if not current_content_source or not current_content_source.strip():
            logger.error("Azure: Content from VIEW_TEXT_WEBSITE_AZURE_CONTENT is empty. Cannot proceed.")
            return azure_data_list

        if current_content_source.startswith('\ufeff'):
            # logger.debug("Removing leading BOM from Azure view_text_website content.")
            current_content_source = current_content_source[1:]

        try:
            soup_vtw = BeautifulSoup(current_content_source, 'lxml-xml')
            items_vtw = soup_vtw.find_all('item')
            if not items_vtw: items_vtw = soup_vtw.find_all('entry')

            entries_to_process_azure = items_vtw
            if limit is not None and limit > 0:
                logger.info(f"Limiting Azure entries (from fallback content) to first {limit}.")
                entries_to_process_azure = items_vtw[:limit]

            if entries_to_process_azure:
                # logger.info(f"Processing {len(entries_to_process_azure)} Azure entries from BeautifulSoup fallback...")
                for i, item_bs in enumerate(entries_to_process_azure):
                    try:
                        # logger.debug(f"Processing Azure fallback entry {i+1}/{len(entries_to_process_azure)}")
                        item_data = {"Title": "N/A", "URL": "N/A", "Date Posted": "N/A", "Description": "N/A", "Links": [], "Status": "N/A", "Type": "N/A", "Products": "N/A", "Categories": "N/A"}
                        title_tag = item_bs.find('title'); item_data["Title"] = title_tag.text.strip() if title_tag else "N/A"
                        link_tag = item_bs.find('link'); item_data["URL"] = link_tag.get('href') if link_tag and link_tag.get('href') else (link_tag.text.strip() if link_tag and link_tag.text else "N/A")
                        pub_date_tag = item_bs.find('pubDate') or item_bs.find('published') or item_bs.find('updated'); published_date_str = pub_date_tag.text.strip() if pub_date_tag else "N/A"
                        item_data["Date Posted"] = published_date_str
                        if "[launched]" in item_data["Title"].lower() or "generally available" in item_data["Title"].lower(): item_data["Status"] = "Launched / Generally Available"
                        elif "[in preview]" in item_data["Title"].lower() or "public preview" in item_data["Title"].lower(): item_data["Status"] = "In Preview"
                        elif "retirement" in item_data["Title"].lower(): item_data["Status"] = "Retirement"
                        rss_categories = [cat.text.strip() for cat in item_bs.find_all('category')]; item_data["Categories"] = ", ".join(rss_categories); item_data["Type"] = rss_categories[0] if rss_categories else "N/A"; item_data["Products"] = ", ".join(rss_categories)

                        if item_data["URL"] == "N/A" or not item_data["URL"].startswith("http"):
                            logger.warning(f"Skipping content scraping for Azure item '{item_data['Title']}' due to invalid URL: {item_data['URL']}")
                            azure_data_list.append(item_data); continue

                        # logger.debug(f"Fetching page content for Azure item: {item_data['URL']}")
                        page_response = requests.get(item_data["URL"], timeout=15); page_response.raise_for_status()
                        page_soup = BeautifulSoup(page_response.content, 'html.parser')

                        def class_contains_keywords(cls):
                            if not cls: return False; s = " ".join(cls).lower() if isinstance(cls, list) else str(cls).lower(); k = ["content", "article", "body", "post", "update", "detail", "summary"]; return any(key in s for key in k)
                        desc_selectors = [{'name': 'div', 'attrs': {'data-bi-name': 'body'}}, {'name': 'section', 'attrs': {'class': 'section-body'}}, {'name': 'article', 'attrs': {}}, {'name': 'div', 'attrs': {'class': class_contains_keywords}}, {'name': 'section', 'attrs': {'class': class_contains_keywords}}, {'name': 'main', 'attrs': {}}]
                        desc_area = None; found_sel_str = "N/A"
                        for sel in desc_selectors:
                            area = page_soup.find(sel['name'], class_=sel['attrs']['class']) if sel['attrs'].get('class') == class_contains_keywords else page_soup.find(sel['name'], **sel['attrs'])
                            if area and (area.find('p') or len(area.get_text(strip=True)) > 200): desc_area = area; found_sel_str = str(sel); #logger.debug(f"Azure page description area found with: {found_sel_str}");
                            break

                        desc_text = "Could not isolate specific article text."
                        if desc_area:
                            if found_sel_str == "{'name': 'main', 'attrs': {}}":
                                # logger.debug("Azure page selected area is <main>, trying more specific div.")
                                main_content_div = desc_area.find('div', class_=class_contains_keywords) or desc_area.find('div', recursive=False)
                                if main_content_div and (main_content_div.find('p') or len(main_content_div.get_text(strip=True)) > 200): desc_area = main_content_div #logger.debug("Using more specific div within <main> for Azure.")
                                # else: logger.debug("Could not find more specific div in <main> for Azure.")

                            paras = desc_area.find_all('p')
                            if paras: desc_text = "\n".join(p.get_text(separator=' ', strip=True) for p in paras).strip()
                            else: desc_text = desc_area.get_text(separator=' ', strip=True) #logger.debug(f"No <p> in Azure area {found_sel_str}, took all text.")
                            generic_texts = ["limited offering to select customers", "available to all azure customers"];
                            if any(gt in desc_text.lower() for gt in generic_texts) and len(desc_text) < 400:
                                logger.warning(f"Azure description for {item_data['URL']} seems generic; marking as 'Could not isolate'. Text: '{desc_text[:100]}...'")
                                desc_text = "Could not isolate specific article text. Found generic content."
                        else: logger.warning(f"No description area found for Azure URL: {item_data['URL']}")

                        item_data["Description"] = (desc_text[:1020] + '...') if len(desc_text) > 1023 else desc_text
                        links_on_page = []
                        if desc_area:
                           for a_tag in desc_area.find_all('a', href=True):
                                href = a_tag['href']
                                if href.startswith("#") or len(href) < 5: continue
                                links_on_page.append(href)
                        item_data["Links"] = ", ".join(links_on_page[:10])
                    except requests.exceptions.RequestException as e_page_req:
                        logger.error(f"Request error for Azure page {item_data.get('URL', 'N/A')}: {e_page_req}")
                        item_data["Description"] = f"Error fetching page: {e_page_req}"
                    except Exception as e_page_scrape:
                        logger.error(f"Unexpected error scraping Azure page {item_data.get('URL', 'N/A')}: {e_page_scrape}", exc_info=True)
                        item_data["Description"] = f"Error scraping page: {e_page_scrape}"
                    azure_data_list.append(item_data)
            else: logger.warning("Azure: BeautifulSoup (on view_text_website content): No <item> or <entry> tags found.")
        except Exception as e_vtw_bs: logger.error(f"Error during BeautifulSoup parsing of Azure view_text_website content: {e_vtw_bs}", exc_info=True)

    logger.info(f"Finished processing {len(azure_data_list)} Azure entries.")
    return azure_data_list

def prepare_workbook(filename="cloud_updates.xlsx"):
    logger.info(f"Preparing workbook: {filename}")
    workbook = Workbook()
    if "Sheet" in workbook.sheetnames: workbook.remove(workbook["Sheet"])
    aws_sheet = workbook.create_sheet("AWS Updates")
    aws_headers = ["Title", "URL", "Date Posted", "Description", "Links", "Product Name"]
    aws_sheet.append(aws_headers)
    azure_sheet = workbook.create_sheet("Azure Updates")
    azure_headers = ["Title", "URL", "Date Posted", "Description", "Links", "Status", "Type", "Products", "Categories"]
    azure_sheet.append(azure_headers)
    try: workbook.save(filename); logger.info(f"Workbook '{filename}' prepared with AWS and Azure sheets.")
    except Exception as e: logger.error(f"Error saving workbook {filename} during preparation: {e}", exc_info=True)
    return workbook, {"aws_sheet_name": "AWS Updates", "azure_sheet_name": "Azure Updates", "aws_headers": aws_headers, "azure_headers": azure_headers}

def write_data_to_excel(workbook, sheet_name, headers, data_list, filename="cloud_updates.xlsx"):
    if not data_list: logger.info(f"No data to write for sheet: {sheet_name}"); return
    logger.info(f"Writing {len(data_list)} items to sheet '{sheet_name}' in '{filename}'.")
    try:
        ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
        if ws.max_row == 1: # Simplified header check
            current_headers = [cell.value for cell in ws[1]]
            if current_headers != headers: ws.append(headers) # Should only happen if sheet was not prepared
        for item_dict in data_list:
            row_values = [str(item_dict.get(header, ""))[:32767] for header in headers] # Excel cell char limit
            ws.append(row_values)
        workbook.save(filename)
        logger.info(f"Successfully wrote data to sheet '{sheet_name}'.")
    except Exception as e: logger.error(f"Error writing to or saving Excel file {filename} for sheet {sheet_name}: {e}", exc_info=True)

def main(args):
    log_level = logging.INFO
    log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'

    # Basic configuration for root logger (if using global logger directly)
    # logging.basicConfig(level=log_level, format=log_format)

    # Specific logger configuration
    logger.setLevel(log_level)

    # Clear existing handlers for this logger to avoid duplicates if main is called multiple times
    if logger.hasHandlers():
        logger.handlers.clear()

    console_handler = logging.StreamHandler()
    console_handler.setFormatter(logging.Formatter(log_format))
    logger.addHandler(console_handler)

    try:
        file_handler = logging.FileHandler("scraper.log", mode='a') # Append mode
        file_handler.setFormatter(logging.Formatter(log_format))
        logger.addHandler(file_handler)
    except Exception as e: # Catch potential errors like permission issues for file handler
        logger.error(f"Could not set up file handler for logging: {e}", exc_info=True)


    logger.info(f"Script started with arguments: {args}")
    excel_filename = args.output

    try:
        workbook, sheet_info = prepare_workbook(excel_filename)

        run_aws = not args.azure_only
        run_azure = not args.aws_only

        if run_aws:
            aws_updates = scrape_aws_updates(limit=args.limit if args.limit and args.limit > 0 else None)
            if aws_updates:
                write_data_to_excel(workbook, sheet_info["aws_sheet_name"], sheet_info["aws_headers"], aws_updates, excel_filename)
            else:
                logger.info("No AWS updates found or error during AWS scraping.")

        if run_azure:
            azure_updates = scrape_azure_updates(limit=args.limit if args.limit and args.limit > 0 else None)
            if azure_updates:
                write_data_to_excel(workbook, sheet_info["azure_sheet_name"], sheet_info["azure_headers"], azure_updates, excel_filename)
            else:
                logger.info("No Azure updates found or error during Azure scraping.")

        logger.info(f"Excel file '{excel_filename}' processing complete.")
    except Exception as e_main:
        logger.critical(f"A critical error occurred in main execution: {e_main}", exc_info=True)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scrape AWS and Azure updates and save to Excel.")
    parser.add_argument("-o", "--output", default="cloud_updates.xlsx", help="Output Excel filename (default: cloud_updates.xlsx)")
    parser.add_argument("--aws-only", action="store_true", help="Only scrape AWS updates")
    parser.add_argument("--azure-only", action="store_true", help="Only scrape Azure updates")
    parser.add_argument("--limit", type=int, default=0, help="Limit the number of entries processed per feed (0 or no value means no limit)")

    parsed_args = parser.parse_args()
    main(parsed_args)
