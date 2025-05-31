import unittest
from unittest.mock import patch, MagicMock, mock_open
import os
import openpyxl
import logging
import requests # Added: For requests.exceptions.RequestException
import argparse # Added: For argparse.Namespace

# Attempt to import functions from src.scraper
# This assumes the agent's execution environment handles PYTHONPATH correctly
# or that tests are run in a way that src is discoverable (e.g. python -m unittest tests.test_scraper from root)
try:
    from src import scraper as scraper_module
except ImportError:
    # Fallback for environments where src might not be directly in PYTHONPATH
    # This is a common pattern but might need adjustment based on execution context
    import sys
    sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
    from src import scraper as scraper_module

# Sample AWS RSS content
SAMPLE_AWS_RSS_XML = """<?xml version="1.0" encoding="UTF-8"?>
<rss version="2.0">
<channel>
  <title>AWS What's New</title>
  <item>
    <title>AWS Test Service 1 Update</title>
    <link>http://aws.amazon.com/test-service-1-update</link>
    <pubDate>Mon, 01 Jan 2024 12:00:00 GMT</pubDate>
    <description>Description for AWS Test Service 1.</description>
  </item>
  <item>
    <title>Amazon EC2 Test Update 2</title>
    <link>http://aws.amazon.com/ec2-test-update-2</link>
    <pubDate>Tue, 02 Jan 2024 12:00:00 GMT</pubDate>
    <description>Description for EC2 Test Update 2.</description>
  </item>
</channel>
</rss>"""

# Sample AWS HTML content for an individual update page
SAMPLE_AWS_PAGE_HTML = """
<html><body>
  <div class="aws-text-box">
    <h1>AWS Test Service 1 Update</h1>
    <p>This is the main content of the update for AWS Test Service 1.</p>
    <p>It includes several <a href="http://example.com/link1">important links</a> and details.</p>
    <a href="http://example.com/another-link">Another link</a>
  </div>
</body></html>
"""

# Sample Azure RSS-like content (similar to VIEW_TEXT_WEBSITE_AZURE_CONTENT)
SAMPLE_AZURE_RSS_XML_STR = """<?xml version="1.0" encoding="utf-8"?>
<rss xmlns:a10="http://www.w3.org/2005/Atom" version="2.0">
<channel>
  <title>Azure service updates</title>
  <item>
    <title>[Launched] Azure Test Service 1 GA</title>
    <link>http://azure.microsoft.com/updates/azure-test-service-1-ga</link>
    <pubDate>Wed, 03 Jan 2024 10:00:00 Z</pubDate>
    <category>AI + machine learning</category>
    <category>Compute</category>
    <description>Description for Azure Test Service 1 GA.</description>
  </item>
  <item>
    <title>Retirement: Old Azure Feature</title>
    <link>http://azure.microsoft.com/updates/old-azure-feature-retirement</link>
    <pubDate>Thu, 04 Jan 2024 11:00:00 Z</pubDate>
    <category>Other</category>
    <category>Retirements</category>
    <description>Old Azure Feature is being retired.</description>
  </item>
</channel>
</rss>"""

# Sample Azure HTML content for an individual update page
SAMPLE_AZURE_PAGE_HTML = """
<html><body>
  <div data-bi-name="body">
    <h2>Azure Test Service 1 GA</h2>
    <p>This is the detailed description of the Azure Test Service 1 GA.</p>
    <p>It has <a href="http://example.com/azure-link1">one link</a> and <a href="http://example.com/azure-link2">another link</a>.</p>
    <div class="tags">
        <a href="/updates/category/ai-machine-learning">AI + Machine Learning</a>
        <a href="/updates/category/compute">Compute</a>
    </div>
  </div>
</body></html>
"""

class MockResponse:
    def __init__(self, content, status_code=200, text=None, apparent_encoding=None):
        self.content = content
        self.text = text if text is not None else content.decode(apparent_encoding or 'utf-8')
        self.status_code = status_code
        self.apparent_encoding = apparent_encoding or 'utf-8'

    def raise_for_status(self):
        if self.status_code >= 400:
            raise requests.exceptions.HTTPError(f"Mock HTTP Error {self.status_code}")

    def json(self): # If any code tries to parse JSON
        import json
        return json.loads(self.text)


class TestScraperBase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        # Suppress logging from the scraper module to keep test output clean
        logging.getLogger('src.scraper').setLevel(logging.CRITICAL + 1) # Disable all logs CRITICAL and below
        # Also for the test's own logger if needed, though usually not necessary
        logging.getLogger(__name__).setLevel(logging.CRITICAL + 1)


class TestAwsScraper(TestScraperBase):
    @patch('src.scraper.requests.get')
    def test_scrape_aws_updates_rss_parsing(self, mock_get):
        # Mock the main RSS feed fetch
        mock_get.return_value = MockResponse(SAMPLE_AWS_RSS_XML.encode('utf-8'))

        # For this test, we only care about RSS parsing, so individual page fetches can return minimal valid HTML
        # or be configured to raise an error if we want to test that path (later)
        # Here, let's make them return a simple page to avoid errors during page scraping part.
        mock_get.side_effect = [
            MockResponse(SAMPLE_AWS_RSS_XML.encode('utf-8')), # First call for RSS
            MockResponse(SAMPLE_AWS_PAGE_HTML.encode('utf-8')), # Call for item 1 page
            MockResponse(SAMPLE_AWS_PAGE_HTML.encode('utf-8'))  # Call for item 2 page
        ]

        aws_data = scraper_module.scrape_aws_updates(limit=2)
        self.assertEqual(len(aws_data), 2)
        self.assertEqual(aws_data[0]['Title'], "AWS Test Service 1 Update")
        self.assertEqual(aws_data[0]['URL'], "http://aws.amazon.com/test-service-1-update")
        self.assertEqual(aws_data[0]['Date Posted'], "2024-01-01")
        self.assertIn("AWS Test", aws_data[0]['Product Name']) # Adjusted expectation

        self.assertEqual(aws_data[1]['Title'], "Amazon EC2 Test Update 2")
        self.assertEqual(aws_data[1]['URL'], "http://aws.amazon.com/ec2-test-update-2")
        self.assertEqual(aws_data[1]['Date Posted'], "2024-01-02")
        self.assertIn("EC2", aws_data[1]['Product Name'])


    @patch('src.scraper.requests.get')
    def test_scrape_aws_updates_page_scraping(self, mock_get):
        # Mock RSS feed to return one item
        single_item_rss = """<?xml version="1.0" encoding="UTF-8"?>
        <rss version="2.0"><channel><item>
            <title>AWS Test Service 1 Update</title>
            <link>http://aws.amazon.com/test-service-1-update</link>
            <pubDate>Mon, 01 Jan 2024 12:00:00 GMT</pubDate>
        </item></channel></rss>"""

        # First call to mock_get is for the RSS feed
        # Second call is for the individual page
        mock_get.side_effect = [
            MockResponse(single_item_rss.encode('utf-8')),
            MockResponse(SAMPLE_AWS_PAGE_HTML.encode('utf-8'))
        ]

        aws_data = scraper_module.scrape_aws_updates(limit=1)
        self.assertEqual(len(aws_data), 1)
        item = aws_data[0]
        self.assertEqual(item['Title'], "AWS Test Service 1 Update")
        self.assertIn("This is the main content", item['Description'])
        self.assertIn("http://example.com/link1", item['Links'])
        self.assertIn("http://example.com/another-link", item['Links'])
        self.assertIn("AWS Test", item['Product Name']) # Adjusted expectation

    @patch('src.scraper.requests.get')
    def test_scrape_aws_network_error_rss(self, mock_get):
        mock_get.side_effect = requests.exceptions.RequestException("Test network error for RSS")
        aws_data = scraper_module.scrape_aws_updates(limit=1)
        self.assertEqual(len(aws_data), 0) # Should return empty list

    @patch('src.scraper.requests.get')
    def test_scrape_aws_network_error_page(self, mock_get):
        # Mock RSS feed to return one item
        single_item_rss = """<?xml version="1.0" encoding="UTF-8"?>
        <rss version="2.0"><channel><item>
            <title>AWS Test Page Network Error</title>
            <link>http://aws.amazon.com/test-page-network-error</link>
            <pubDate>Mon, 01 Jan 2024 12:00:00 GMT</pubDate>
        </item></channel></rss>"""

        mock_get.side_effect = [
            MockResponse(single_item_rss.encode('utf-8')), # For RSS
            requests.exceptions.RequestException("Test network error for page") # For page fetch
        ]
        aws_data = scraper_module.scrape_aws_updates(limit=1)
        self.assertEqual(len(aws_data), 1)
        self.assertIn("Error fetching page", aws_data[0]['Description'])


class TestAzureScraper(TestScraperBase):
    @patch('src.scraper.requests.get') # This mock will handle ALL calls to requests.get in scrape_azure_updates
    def test_scrape_azure_updates_fallback_parsing(self, mock_requests_get_azure):

        def selective_mock_get(url, timeout=None):
            if url == scraper_module.AZURE_RSS_URL:
                # This is the initial call to get the main Azure RSS feed.
                # Return content that feedparser will be mocked to fail on.
                return MockResponse("Simulated problematic Azure RSS for feedparser".encode('utf-8'))
            elif "azure.microsoft.com/updates/" in url: # Matches page URLs from SAMPLE_AZURE_RSS_XML_STR
                return MockResponse(SAMPLE_AZURE_PAGE_HTML.encode('utf-8'))
            # Fallback for any other unexpected URL
            return MockResponse("Default unexpected mock response".encode('utf-8'), status_code=404)

        mock_requests_get_azure.side_effect = selective_mock_get

        with patch.object(scraper_module, 'VIEW_TEXT_WEBSITE_AZURE_CONTENT', SAMPLE_AZURE_RSS_XML_STR):
            # Mock feedparser to ensure it fails on the (mocked) initial RSS response,
            # forcing the use of VIEW_TEXT_WEBSITE_AZURE_CONTENT
            with patch('src.scraper.feedparser.parse', MagicMock(return_value=MagicMock(bozo=1, entries=[]))):
                azure_data = scraper_module.scrape_azure_updates(limit=2)

        self.assertEqual(len(azure_data), 2)
        self.assertEqual(azure_data[0]['Title'], "[Launched] Azure Test Service 1 GA")
        self.assertEqual(azure_data[0]['URL'], "http://azure.microsoft.com/updates/azure-test-service-1-ga")
        self.assertEqual(azure_data[0]['Date Posted'], "Wed, 03 Jan 2024 10:00:00 Z") # Raw date
        self.assertEqual(azure_data[0]['Status'], "Launched / Generally Available")
        self.assertIn("AI + machine learning", azure_data[0]['Categories'])
        self.assertIn("Compute", azure_data[0]['Categories'])

        # Test page scraping details for the first item
        self.assertIn("detailed description of the Azure Test Service 1 GA", azure_data[0]['Description'])
        self.assertIn("http://example.com/azure-link1", azure_data[0]['Links'])


class TestExcelFunctions(TestScraperBase):
    TEST_EXCEL_FILENAME = "test_cloud_updates.xlsx"

    def tearDown(self):
        if os.path.exists(self.TEST_EXCEL_FILENAME):
            os.remove(self.TEST_EXCEL_FILENAME)

    def test_prepare_workbook(self):
        workbook, sheet_info = scraper_module.prepare_workbook(filename=self.TEST_EXCEL_FILENAME)
        self.assertIsNotNone(workbook)
        self.assertTrue(os.path.exists(self.TEST_EXCEL_FILENAME))

        loaded_workbook = openpyxl.load_workbook(self.TEST_EXCEL_FILENAME)
        self.assertIn("AWS Updates", loaded_workbook.sheetnames)
        self.assertIn("Azure Updates", loaded_workbook.sheetnames)

        aws_sheet = loaded_workbook["AWS Updates"]
        self.assertEqual(aws_sheet['A1'].value, "Title")
        self.assertEqual(aws_sheet['F1'].value, "Product Name")

        azure_sheet = loaded_workbook["Azure Updates"]
        self.assertEqual(azure_sheet['A1'].value, "Title")
        self.assertEqual(azure_sheet['I1'].value, "Categories")

    def test_write_data_to_excel(self):
        workbook, sheet_info = scraper_module.prepare_workbook(filename=self.TEST_EXCEL_FILENAME)

        aws_sample_data = [
            {"Title": "AWS Test 1", "URL": "url1", "Date Posted": "2024-01-01", "Description": "Desc1", "Links": "link1", "Product Name": "EC2"},
        ]
        scraper_module.write_data_to_excel(workbook, sheet_info["aws_sheet_name"], sheet_info["aws_headers"], aws_sample_data, self.TEST_EXCEL_FILENAME)

        azure_sample_data = [
            {"Title": "Azure Test 1", "URL": "url_az1", "Date Posted": "2024-01-02", "Description": "AzDesc1", "Links": "azlink1", "Status": "GA", "Type": "Features", "Products": "Compute", "Categories": "Compute"},
        ]
        scraper_module.write_data_to_excel(workbook, sheet_info["azure_sheet_name"], sheet_info["azure_headers"], azure_sample_data, self.TEST_EXCEL_FILENAME)

        loaded_workbook = openpyxl.load_workbook(self.TEST_EXCEL_FILENAME)
        aws_sheet = loaded_workbook["AWS Updates"]
        self.assertEqual(aws_sheet['A2'].value, "AWS Test 1")
        self.assertEqual(aws_sheet['F2'].value, "EC2")

        azure_sheet = loaded_workbook["Azure Updates"]
        self.assertEqual(azure_sheet['A2'].value, "Azure Test 1")
        self.assertEqual(azure_sheet['G2'].value, "Features")


class TestMainFunctionLogic(TestScraperBase):
    TEST_OUTPUT_XLSX = "main_test_output.xlsx"

    def tearDown(self):
        if os.path.exists(self.TEST_OUTPUT_XLSX):
            os.remove(self.TEST_OUTPUT_XLSX)
        if os.path.exists("scraper.log"): # Clean up log file if created by test run
            os.remove("scraper.log")


    @patch('src.scraper.scrape_aws_updates')
    @patch('src.scraper.scrape_azure_updates')
    @patch('src.scraper.prepare_workbook') # Mock prepare to control workbook object
    @patch('src.scraper.write_data_to_excel') # Mock actual writing
    def test_main_runs_both_scrapers_default(self, mock_write_excel, mock_prepare_workbook, mock_scrape_azure, mock_scrape_aws):
        # Setup mock for prepare_workbook
        mock_wb = openpyxl.Workbook()
        mock_sheet_info = {
            "aws_sheet_name": "AWS Updates", "azure_sheet_name": "Azure Updates",
            "aws_headers": ["Title", "URL", "Date Posted", "Description", "Links", "Product Name"],
            "azure_headers": ["Title", "URL", "Date Posted", "Description", "Links", "Status", "Type", "Products", "Categories"]
        }
        mock_prepare_workbook.return_value = (mock_wb, mock_sheet_info)

        mock_scrape_aws.return_value = [{"Title": "AWS Data"}] # Dummy data
        mock_scrape_azure.return_value = [{"Title": "Azure Data"}] # Dummy data

        args = argparse.Namespace(output=self.TEST_OUTPUT_XLSX, aws_only=False, azure_only=False, limit=0)
        scraper_module.main(args)

        mock_scrape_aws.assert_called_once_with(limit=None)
        mock_scrape_azure.assert_called_once_with(limit=None)
        self.assertEqual(mock_write_excel.call_count, 2)
        # Check if scraper.log was created (logging is initialized in main)
        # self.assertTrue(os.path.exists("scraper.log"))


    @patch('src.scraper.scrape_aws_updates')
    @patch('src.scraper.scrape_azure_updates')
    @patch('src.scraper.prepare_workbook')
    @patch('src.scraper.write_data_to_excel')
    def test_main_runs_aws_only(self, mock_write_excel, mock_prepare_workbook, mock_scrape_azure, mock_scrape_aws):
        mock_wb = openpyxl.Workbook()
        mock_sheet_info = {"aws_sheet_name": "AWS Updates", "aws_headers": []} # Simplified
        mock_prepare_workbook.return_value = (mock_wb, mock_sheet_info)
        mock_scrape_aws.return_value = [{"Title": "AWS Data"}]

        args = argparse.Namespace(output=self.TEST_OUTPUT_XLSX, aws_only=True, azure_only=False, limit=5)
        scraper_module.main(args)

        mock_scrape_aws.assert_called_once_with(limit=5)
        mock_scrape_azure.assert_not_called()
        mock_write_excel.assert_called_once() # Only for AWS

    @patch('src.scraper.scrape_aws_updates')
    @patch('src.scraper.scrape_azure_updates')
    @patch('src.scraper.prepare_workbook')
    @patch('src.scraper.write_data_to_excel')
    def test_main_runs_azure_only_with_limit(self, mock_write_excel, mock_prepare_workbook, mock_scrape_azure, mock_scrape_aws):
        mock_wb = openpyxl.Workbook()
        mock_sheet_info = {"azure_sheet_name": "Azure Updates", "azure_headers": []} # Simplified
        mock_prepare_workbook.return_value = (mock_wb, mock_sheet_info)
        mock_scrape_azure.return_value = [{"Title": "Azure Data"}]

        args = argparse.Namespace(output=self.TEST_OUTPUT_XLSX, aws_only=False, azure_only=True, limit=2)
        scraper_module.main(args)

        mock_scrape_aws.assert_not_called()
        mock_scrape_azure.assert_called_once_with(limit=2)
        mock_write_excel.assert_called_once() # Only for Azure

if __name__ == '__main__':
    unittest.main()
